#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
主界面 - 通道查询、编辑、批量修改、Excel导入导出
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import hashlib
import requests
import threading
from datetime import datetime
import os

# 尝试导入 openpyxl（非必需，但导出/导入需要）
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from progress_dialog import ProgressDialog

DEFAULT_DEVICE_ID = "35021139492007094538"


class MainApplication:
    def __init__(self, root, token, user_info, host):
        self.root = root
        self.access_token = token
        self.login_user = user_info
        self.server_host = tk.StringVar(value=host)

        self.root.title("WVP通道管理工具 v6.0 by Xiaoabiao")
        self.root.geometry("1000x750")
        self.root.minsize(900, 600)

        self.device_id_var = tk.StringVar(value=DEFAULT_DEVICE_ID)
        self.status_text = tk.StringVar(value=f"已登录: {user_info.get('username', '')}")
        self.page_num = 1
        self.page_size = 100
        self.total_channels = 0
        self.all_channels = []
        self.item_to_channel = {}

        self.edit_entry = None
        self.edit_item = None
        self.edit_column = None

        self.check_vars = {}
        self.select_all_var = tk.BooleanVar(value=False)

        self.progress = None          # 进度弹窗实例

        self.setup_styles()
        self.build_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", rowheight=26, font=("Microsoft YaHei", 9))
        style.configure("Treeview.Heading", font=("Microsoft YaHei", 10, "bold"))
        style.configure("TButton", font=("Microsoft YaHei", 9))
        style.configure("TLabel", font=("Microsoft YaHei", 9))
        style.configure("TEntry", font=("Microsoft YaHei", 9))

    def build_ui(self):
        # 顶栏
        top_bar = ttk.Frame(self.root)
        top_bar.pack(fill=tk.X, padx=10, pady=(5, 0))
        ttk.Label(top_bar, text=f"👤 {self.login_user.get('username', '')}",
                  font=("Microsoft YaHei", 10, "bold")).pack(side=tk.LEFT)
        ttk.Button(top_bar, text="退出登录", command=self.logout).pack(side=tk.RIGHT, padx=5)
        ttk.Separator(self.root, orient='horizontal').pack(fill=tk.X, padx=10, pady=(2, 0))

        # 查询区域
        query_frame = ttk.LabelFrame(self.root, text=" 设备通道查询 ", padding=10)
        query_frame.pack(fill=tk.X, padx=10, pady=5)

        qrow = ttk.Frame(query_frame)
        qrow.pack(fill=tk.X)
        ttk.Label(qrow, text="设备国标ID:").pack(side=tk.LEFT, padx=(0, 5))
        self.device_entry = ttk.Entry(qrow, textvariable=self.device_id_var, width=30)
        self.device_entry.pack(side=tk.LEFT, padx=(0, 10))
        self.query_btn = ttk.Button(qrow, text="🔍 查询通道", command=self.do_query_channels, width=12)
        self.query_btn.pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(qrow, text="🔄 刷新", command=self.do_refresh, width=8).pack(side=tk.LEFT)

        self.export_btn = ttk.Button(qrow, text="📤 导出Excel", command=self.export_excel, width=12, state=tk.DISABLED)
        self.export_btn.pack(side=tk.LEFT, padx=(20, 5))
        self.import_btn = ttk.Button(qrow, text="📥 导入Excel", command=self.import_excel, width=12, state=tk.DISABLED)
        self.import_btn.pack(side=tk.LEFT)

        self.page_info_var = tk.StringVar(value="")
        ttk.Label(qrow, textvariable=self.page_info_var, foreground="gray").pack(side=tk.RIGHT, padx=5)

        # 表格区域
        table_frame = ttk.LabelFrame(self.root, text=" 通道列表 (点击☐选择，双击名称/区域编码编辑) ", padding=5)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        table_frame.grid_rowconfigure(0, weight=0)
        table_frame.grid_rowconfigure(1, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        batch_row = ttk.Frame(table_frame)
        batch_row.grid(row=0, column=0, sticky="ew", pady=(0, 5))

        self.select_all_cb = ttk.Checkbutton(batch_row, text="全选/取消全选", variable=self.select_all_var,
                                             command=self.toggle_select_all)
        self.select_all_cb.pack(side=tk.LEFT, padx=(5, 15))

        self.batch_region_btn = ttk.Button(batch_row, text="📝 批量修改区域编码", command=self.batch_modify_region,
                                           state=tk.DISABLED)
        self.batch_region_btn.pack(side=tk.LEFT, padx=5)

        tree_container = ttk.Frame(table_frame)
        tree_container.grid(row=1, column=0, sticky="nsew")
        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)

        columns = ("选择", "序号", "设备ID", "名称", "通道类型", "父设备", "区域编码", "数据库ID")
        self.tree = ttk.Treeview(tree_container, columns=columns, show="headings", selectmode="browse")

        col_widths = {"选择": 40, "序号": 45, "设备ID": 180, "名称": 200, "通道类型": 80,
                      "父设备": 120, "区域编码": 100, "数据库ID": 80}
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=col_widths.get(col, 100), anchor=tk.CENTER, minwidth=40)

        vsb = ttk.Scrollbar(tree_container, orient=tk.VERTICAL, command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_container, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        self.tree.bind("<Button-1>", self.on_checkbox_click)
        self.tree.bind("<Double-1>", self.on_double_click)

        self.statusbar_var = tk.StringVar(value="就绪")
        statusbar = ttk.Label(self.root, textvariable=self.statusbar_var, relief=tk.SUNKEN,
                              anchor=tk.W, font=("Microsoft YaHei", 8), padding=(5, 2))
        statusbar.pack(side=tk.BOTTOM, fill=tk.X)

        self.update_ui_state()

    # ---------- UI 状态 ----------
    def update_ui_state(self):
        logged_in = self.access_token is not None
        has_channels = len(self.all_channels) > 0
        self.query_btn.configure(state=tk.NORMAL if logged_in else tk.DISABLED)
        self.export_btn.configure(state=tk.NORMAL if (logged_in and has_channels) else tk.DISABLED)
        self.import_btn.configure(state=tk.NORMAL if logged_in else tk.DISABLED)
        self.batch_region_btn.configure(state=tk.NORMAL if (logged_in and has_channels) else tk.DISABLED)

    def set_statusbar(self, msg):
        self.statusbar_var.set(f"{msg}  [{datetime.now().strftime('%H:%M:%S')}]")

    # ---------- 查询 ----------
    def do_query_channels(self):
        if not self.access_token:
            messagebox.showwarning("警告", "请先登录")
            return
        device_id = self.device_id_var.get().strip()
        if not device_id:
            messagebox.showwarning("警告", "请输入设备国标ID")
            return
        self.set_statusbar(f"正在查询设备 {device_id} 的全部通道...")
        self.query_btn.configure(state=tk.DISABLED, text="查询中...")
        self.cancel_edit()

        def task():
            try:
                host = self.server_host.get().strip().rstrip('/')
                url = f"{host}/api/device/query/devices/{device_id}/channels"
                headers = {
                    "Accept": "*/*",
                    "access-token": self.access_token,
                    "Content-Type": "application/x-www-form-urlencoded"
                }
                all_channels = []
                page = 1
                count = self.page_size       # 每页请求条数，默认100
                total = 0

                while True:
                    params = {
                        "page": page,
                        "count": count,
                        "query": "",
                        "cameraQuery": "",
                        "nvrQuery": ""
                    }
                    resp = requests.get(url, headers=headers, params=params, timeout=30)

                    if resp.status_code == 200:
                        data = resp.json()
                        if data.get("code") != 0:
                            self.root.after(0, lambda m=data.get("msg", "未知错误"): self._query_fail(m))
                            return

                        inner = data.get("data", {})
                        channels = inner.get("list", [])
                        total = inner.get("total", 0)
                        all_channels.extend(channels)

                        # 更新界面状态提示
                        self.root.after(0, lambda p=page: self.set_statusbar(f"正在加载第 {p} 页..."))

                        # 判断是否已取完所有数据
                        if len(channels) < count or len(all_channels) >= total:
                            break
                        page += 1
                    else:
                        msg = resp.json().get("msg", f"HTTP {resp.status_code}")
                        self.root.after(0, lambda m=msg: self._query_fail(m))
                        return

                self.all_channels = all_channels
                self.total_channels = total
                self.root.after(0, lambda: self._query_success(all_channels, total, device_id))

            except Exception as e:
                self.root.after(0, lambda e=e: self._query_fail(str(e)))

        threading.Thread(target=task, daemon=True).start()

    def _query_success(self, channels, total, device_id):
        self.query_btn.configure(state=tk.NORMAL, text="🔍 查询通道")
        self.tree.delete(*self.tree.get_children())
        self.item_to_channel.clear()
        self.check_vars.clear()
        self.select_all_var.set(False)

        for i, ch in enumerate(channels, 1):
            db_id = ch.get("id")
            ch_type = ch.get("channelType", 0)
            type_text = "子目录" if ch_type else "设备通道"
            var = tk.BooleanVar(value=False)
            values = ("☐", i, ch.get("deviceId", ""), ch.get("name", ""),
                      type_text, ch.get("parentId", "") or "", ch.get("civilCode", ""), db_id)
            item = self.tree.insert("", tk.END, values=values)
            self.item_to_channel[item] = ch
            self.check_vars[item] = var
            var.trace_add("write", lambda *args, it=item: self.update_check_display(it))

        # 显示总条数（无分页提示）
        self.page_info_var.set(f"共 {total} 条")
        self.set_statusbar(f"查询成功 - 设备 {device_id} 下共 {total} 个通道")
        self.update_ui_state()

    def _query_fail(self, msg):
        self.query_btn.configure(state=tk.NORMAL, text="🔍 查询通道")
        self.set_statusbar(f"查询失败: {msg}")
        messagebox.showerror("查询失败", msg)

    def do_refresh(self):
        if self.device_id_var.get().strip():
            self.do_query_channels()

    # ---------- 复选框 ----------
    def update_check_display(self, item):
        var = self.check_vars.get(item)
        if var:
            self.tree.set(item, "#1", "☑" if var.get() else "☐")

    def on_checkbox_click(self, event):
        if self.tree.identify_column(event.x) != "#1":
            return
        item = self.tree.identify_row(event.y)
        if not item:
            return
        var = self.check_vars.get(item)
        if var:
            var.set(not var.get())
            self.update_select_all_state()

    def toggle_select_all(self):
        state = self.select_all_var.get()
        for var in self.check_vars.values():
            var.set(state)

    def update_select_all_state(self):
        if not self.check_vars:
            self.select_all_var.set(False)
            return
        all_checked = all(v.get() for v in self.check_vars.values())
        self.select_all_var.set(all_checked)

    def get_selected_channels(self):
        return [self.item_to_channel[it] for it, var in self.check_vars.items()
                if var.get() and it in self.item_to_channel]

    # ---------- 批量修改区域编码 ----------
    def batch_modify_region(self):
        selected = self.get_selected_channels()
        if not selected:
            messagebox.showwarning("提示", "请至少勾选一个通道")
            return
        new_region = simpledialog.askstring("批量修改区域编码",
                                            f"已选中 {len(selected)} 个通道\n请输入新的区域编码:",
                                            parent=self.root)
        if not new_region or not new_region.strip():
            return
        new_region = new_region.strip()
        if not messagebox.askyesno("确认", f"将为选中的 {len(selected)} 个通道设置区域编码为:\n'{new_region}'?\n\n确认修改?"):
            return
        self.batch_region_btn.configure(state=tk.DISABLED)
        self.set_statusbar("正在批量修改区域编码...")
        threading.Thread(target=self._batch_update_region, args=(selected, new_region), daemon=True).start()

    def _batch_update_region(self, channels, new_region):
        success = fail = 0
        host = self.server_host.get().strip().rstrip('/')
        headers = {"Accept": "*/*", "access-token": self.access_token, "Content-Type": "application/json"}
        for idx, ch in enumerate(channels, 1):
            updates = {"gbCivilCode": new_region, "civilCode": new_region}
            try:
                body = self.build_channel_body(ch, updates)
                resp = requests.post(f"{host}/api/common/channel/update", headers=headers, json=body, timeout=15)
                if resp.status_code == 200 and resp.json().get("code") == 0:
                    success += 1
                else:
                    fail += 1
            except:
                fail += 1
            self.root.after(0, lambda c=idx: self.set_statusbar(f"批量修改中: {c}/{len(channels)}"))
        self.root.after(0, lambda: self._batch_finished(success, fail))

    def build_channel_body(self, channel, updates):
        body = {
            "gbId": channel.get("id", 0),
            "gbDeviceId": channel.get("deviceId", ""),
            "gbName": channel.get("name", ""),
            "gbManufacturer": channel.get("gbManufacturer", ""),
            "gbModel": channel.get("gbModel", ""),
            "gbOwner": channel.get("gbOwner", ""),
            "gbCivilCode": channel.get("civilCode", ""),
            "gbBlock": channel.get("gbBlock", ""),
            "gbAddress": channel.get("gbAddress", ""),
            "gbParental": channel.get("gbParental", 0),
            "gbParentId": channel.get("gbParentId", ""),
            "gbSafetyWay": channel.get("gbSafetyWay", 0),
            "gbRegisterWay": channel.get("gbRegisterWay", 0),
            "gbCertNum": channel.get("gbCertNum", ""),
            "gbCertifiable": channel.get("gbCertifiable", 0),
            "gbErrCode": channel.get("gbErrCode", 0),
            "gbEndTime": channel.get("gbEndTime", ""),
            "gbSecrecy": channel.get("gbSecrecy", 0),
            "gbIpAddress": channel.get("gbIpAddress", ""),
            "gbPort": channel.get("gbPort", 0),
            "gbPassword": channel.get("gbPassword", ""),
            "gbStatus": channel.get("gbStatus", ""),
            "gbLongitude": channel.get("gbLongitude", 0),
            "gbLatitude": channel.get("gbLatitude", 0),
            "gpsAltitude": channel.get("gpsAltitude", 0),
            "gpsSpeed": channel.get("gpsSpeed", 0),
            "gpsDirection": channel.get("gpsDirection", 0),
            "gpsTime": channel.get("gpsTime", ""),
            "gbBusinessGroupId": channel.get("gbBusinessGroupId", ""),
            "gbPtzType": channel.get("gbPtzType", 0),
            "gbPositionType": channel.get("gbPositionType", 0),
            "gbRoomType": channel.get("gbRoomType", 0),
            "gbUseType": channel.get("gbUseType", 0),
            "gbSupplyLightType": channel.get("gbSupplyLightType", 0),
            "gbDirectionType": channel.get("gbDirectionType", 0),
            "gbResolution": channel.get("gbResolution", ""),
            "gbDownloadSpeed": channel.get("gbDownloadSpeed", ""),
            "gbSvcSpaceSupportMod": channel.get("gbSvcSpaceSupportMod", 0),
            "gbSvcTimeSupportMode": channel.get("gbSvcTimeSupportMode", 0),
            "recordPLan": channel.get("recordPLan", 0),
            "dataType": channel.get("dataType", 0),
            "dataDeviceId": channel.get("dataDeviceId", 0),
            "createTime": channel.get("createTime", ""),
            "updateTime": channel.get("updateTime", ""),
        }
        body.update(updates)
        return body

    def _batch_finished(self, success, fail):
        self.batch_region_btn.configure(state=tk.NORMAL)
        self.set_statusbar(f"批量修改完成: 成功 {success}, 失败 {fail}")
        messagebox.showinfo("结果", f"成功: {success}\n失败: {fail}")
        if success > 0:
            self.do_refresh()

    # ---------- 双击编辑 ----------
    def on_double_click(self, event):
        col = self.tree.identify_column(event.x)
        if col == "#1":
            return
        if self.edit_entry:
            self.save_edit()
        item = self.tree.identify_row(event.y)
        if not item or col not in ("#4", "#7"):
            return
        self.edit_item = item
        self.edit_column = col
        value = self.tree.set(item, col)
        bbox = self.tree.bbox(item, col)
        if not bbox:
            return
        x, y, w, h = bbox
        self.edit_entry = tk.Entry(self.tree, font=("Microsoft YaHei", 9))
        self.edit_entry.place(x=x, y=y, width=w, height=h)
        self.edit_entry.insert(0, value)
        self.edit_entry.select_range(0, tk.END)
        self.edit_entry.focus_set()
        self.edit_entry.bind("<Return>", lambda e: self.save_edit())
        self.edit_entry.bind("<FocusOut>", self.on_focus_out)

    def on_focus_out(self, event):
        if self.edit_entry:
            self.root.after(100, self.save_edit)

    def save_edit(self):
        if not self.edit_entry:
            return
        new_value = self.edit_entry.get().strip()
        item, col = self.edit_item, self.edit_column
        self.cancel_edit()
        if not item or not new_value:
            return
        channel = self.item_to_channel.get(item)
        if not channel:
            return
        if col == "#4":
            old = channel.get("name", "")
            if new_value == old:
                return
            updates = {"gbName": new_value, "name": new_value}
            field = "名称"
        else:
            old = channel.get("civilCode", "")
            if new_value == old:
                return
            updates = {"gbCivilCode": new_value, "civilCode": new_value}
            field = "区域编码"
        if not messagebox.askyesno("确认", f"将 {field} 从 '{old}' 改为 '{new_value}'?"):
            return
        threading.Thread(target=self.do_update, args=(channel, updates, new_value, item, col), daemon=True).start()

    def cancel_edit(self):
        if self.edit_entry:
            self.edit_entry.destroy()
            self.edit_entry = None
            self.edit_item = None
            self.edit_column = None

    def do_update(self, channel, updates, new_value, item, column):
        try:
            host = self.server_host.get().strip().rstrip('/')
            headers = {"Accept": "*/*", "access-token": self.access_token, "Content-Type": "application/json"}
            body = self.build_channel_body(channel, updates)
            resp = requests.post(f"{host}/api/common/channel/update", headers=headers, json=body, timeout=15)
            if resp.status_code == 200 and resp.json().get("code") == 0:
                self.root.after(0, lambda: self.tree.set(item, column, new_value))
                self.root.after(0, lambda: self.set_statusbar("修改成功"))
            else:
                msg = resp.json().get("msg", f"HTTP {resp.status_code}")
                self.root.after(0, lambda m=msg: messagebox.showerror("修改失败", m))
        except Exception as e:
            self.root.after(0, lambda e=e: messagebox.showerror("修改异常", str(e)))

    # ---------- 导出Excel ----------
    def export_excel(self):
        if not self.all_channels:
            messagebox.showwarning("提示", "通道列表为空")
            return
        if not HAS_OPENPYXL:
            messagebox.showerror("缺少库", "请安装 openpyxl: pip install openpyxl")
            return
        file = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel", "*.xlsx")], title="导出通道列表")
        if not file:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "通道列表"
            headers = ["序号", "设备ID", "名称", "通道类型", "父设备", "区域编码", "数据库ID"]
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            for i, ch in enumerate(self.all_channels, 1):
                row = [i, ch.get("deviceId", ""), ch.get("name", ""),
                       "子目录" if ch.get("channelType") else "设备通道",
                       ch.get("parentId", ""), ch.get("civilCode", ""), ch.get("id", "")]
                for j, v in enumerate(row, 1):
                    ws.cell(row=i+1, column=j, value=v)
            for i, w in enumerate([6, 22, 25, 10, 18, 15, 10], 1):
                ws.column_dimensions[ws.cell(1, i).column_letter].width = w
            wb.save(file)
            self.set_statusbar(f"导出成功: {os.path.basename(file)}")
            messagebox.showinfo("成功", f"已导出到:\n{file}")
        except Exception as e:
            messagebox.showerror("失败", str(e))

    # ---------- 导入Excel（带进度条）----------
    def import_excel(self):
        if not self.access_token:
            messagebox.showwarning("警告", "请先登录")
            return
        if not HAS_OPENPYXL:
            messagebox.showerror("缺少库", "请安装 openpyxl: pip install openpyxl")
            return
        if not self.all_channels:
            messagebox.showwarning("提示", "请先查询通道再导入")
            return
        file = filedialog.askopenfilename(title="选择修改后的Excel", filetypes=[("Excel", "*.xlsx")])
        if not file:
            return
        try:
            wb = load_workbook(file)
            ws = wb.active
        except Exception as e:
            messagebox.showerror("读取失败", str(e))
            return
        excel = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 7:
                continue
            did = row[6]
            if not did:
                continue
            try:
                did = int(did)
            except:
                continue
            excel[did] = {"name": str(row[2] or "").strip(), "civilCode": str(row[5] or "").strip()}
        if not excel:
            messagebox.showwarning("无数据", "Excel无有效数据")
            return
        tasks = []
        for ch in self.all_channels:
            if ch["id"] in excel:
                n, c = excel[ch["id"]]["name"], excel[ch["id"]]["civilCode"]
                if n != ch.get("name", "") or c != ch.get("civilCode", ""):
                    tasks.append((ch, n, c))
        if not tasks:
            messagebox.showinfo("提示", "没有需要修改的数据")
            return
        if not messagebox.askyesno("确认", f"检测到 {len(tasks)} 条修改，是否继续?"):
            return
        self.progress = ProgressDialog(self.root, "正在导入修改", len(tasks))
        self.import_btn.configure(state=tk.DISABLED)
        threading.Thread(target=self.batch_excel_update, args=(tasks,), daemon=True).start()

    def batch_excel_update(self, tasks):
        success = fail = 0
        host = self.server_host.get().strip().rstrip('/')
        headers = {"Accept": "*/*", "access-token": self.access_token, "Content-Type": "application/json"}
        total = len(tasks)
        for i, (ch, nn, nc) in enumerate(tasks, 1):
            if self.progress and self.progress.is_cancelled():
                break
            upd = {}
            if nn != ch.get("name", ""):
                upd["gbName"] = nn
                upd["name"] = nn
            if nc != ch.get("civilCode", ""):
                upd["gbCivilCode"] = nc
                upd["civilCode"] = nc
            try:
                body = self.build_channel_body(ch, upd)
                resp = requests.post(f"{host}/api/common/channel/update", headers=headers, json=body, timeout=15)
                if resp.status_code == 200 and resp.json().get("code") == 0:
                    success += 1
                else:
                    fail += 1
            except:
                fail += 1
            self.root.after(0, lambda c=i, t=total: self.progress.update(c, t, f"正在处理 {c}/{t}"))
        self.root.after(0, self.progress.close)
        self.root.after(0, lambda: self._batch_finished(success, fail))
        self.root.after(0, lambda: self.import_btn.configure(state=tk.NORMAL))

    # ---------- 退出 ----------
    def logout(self):
        if messagebox.askyesno("退出", "确定要退出登录吗？"):
            self.root.destroy()

    def on_close(self):
        self.cancel_edit()
        self.root.destroy()